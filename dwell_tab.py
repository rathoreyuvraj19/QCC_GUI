"""
dwell_tab.py

"Dwell" - the main per-QTRM beam command (Command Type 0x01): each of the
96 QTRMs gets its own 4-channel Control/Tx Phase/Tx Atten/Rx Phase/Rx Atten
set, all sent together in one frame (unlike Cal/Isolation/Soft Reset, Dwell
has no single-QTRM-target convention).

Two ways to fill the 96-row table: type values in directly, or import an
Excel (.xlsx) sheet - one row per QTRM, columns "QTRM ID" (optional, else
row order = QTRM 1-96) and "Ch{1-4} Control/Tx Phase/Tx Atten/Rx Phase/Rx
Atten". Imported data is latched into the table (stays until edited or
re-imported) until Send is pressed.

Every QTRM's Dwell command requests a Link-type status response (per
Yuvraj: every command except Status and Soft Reset does) - the 96-cell LED
matrix (reused from link_test_tab.py) shows which QTRMs acknowledged.
"""

from PySide6.QtCore import QAbstractTableModel, Qt, QModelIndex, Signal
from PySide6.QtWidgets import (
    QAbstractItemView, QFileDialog, QHBoxLayout, QHeaderView, QLabel,
    QMessageBox, QPushButton, QScrollArea, QTableView, QVBoxLayout, QWidget,
)

from header_panel import HeaderPanel
from link_test_tab import LedMatrix, _NOT_LINKED_COLOR, _PENDING_COLOR
from packet import NUM_QTRM, QTRMChannel

PHASE_MAX = 63    # 6-bit phase (frame_type.vhd: No_of_phase_bits = 6)
ATTEN_MAX = 63    # 6-bit attenuation (frame_type.vhd: No_of_Attenuator_bits = 6)

# (label, attribute, min, max)
_CHANNEL_FIELDS = [
    ("Control", "control", 0, 255),
    ("Tx Phase", "tx_phase", 0, PHASE_MAX),
    ("Tx Atten", "tx_atten", 0, ATTEN_MAX),
    ("Rx Phase", "rx_phase", 0, PHASE_MAX),
    ("Rx Atten", "rx_atten", 0, ATTEN_MAX),
]


def _build_columns():
    cols = [("QTRM ID", None, False)]
    for ch in range(1, 5):
        for label, attr, lo, hi in _CHANNEL_FIELDS:
            cols.append((f"Ch{ch} {label}", (ch - 1, attr, lo, hi), True))
    return cols


COLUMNS = _build_columns()


def _default_channels():
    return [[QTRMChannel() for _ in range(4)] for _ in range(NUM_QTRM)]


class DwellTableModel(QAbstractTableModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.channels = _default_channels()

    def rowCount(self, parent=QModelIndex()):
        return NUM_QTRM

    def columnCount(self, parent=QModelIndex()):
        return len(COLUMNS)

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return COLUMNS[section][0]
        return str(section + 1)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        _, key, _ = COLUMNS[index.column()]
        if role in (Qt.DisplayRole, Qt.EditRole):
            if key is None:
                return index.row() + 1
            ch_idx, attr, _, _ = key
            return getattr(self.channels[index.row()][ch_idx], attr)
        return None

    def setData(self, index, value, role=Qt.EditRole):
        if role != Qt.EditRole or not index.isValid():
            return False
        _, key, editable = COLUMNS[index.column()]
        if not editable:
            return False
        try:
            ivalue = int(value)
        except (TypeError, ValueError):
            return False
        ch_idx, attr, lo, hi = key
        ivalue = max(lo, min(hi, ivalue))
        setattr(self.channels[index.row()][ch_idx], attr, ivalue)
        self.dataChanged.emit(index, index, [role])
        return True

    def flags(self, index):
        if not index.isValid():
            return Qt.NoItemFlags
        _, _, editable = COLUMNS[index.column()]
        base = Qt.ItemIsEnabled | Qt.ItemIsSelectable
        if editable:
            base |= Qt.ItemIsEditable
        return base

    def load_channels(self, channels):
        self.beginResetModel()
        self.channels = channels
        self.endResetModel()

    def get_channels(self):
        return self.channels


def _clamp(value, lo, hi):
    try:
        ivalue = int(value)
    except (TypeError, ValueError):
        return lo
    return max(lo, min(hi, ivalue))


def load_channels_from_excel(path: str):
    """
    Parse an .xlsx sheet into a NUM_QTRM-length list of 4 QTRMChannel each.
    First row is headers; "QTRM ID" (1-based) is optional - if absent, row
    order is taken as QTRM 1..96. Missing Ch{n} columns default to 0.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Spreadsheet has no rows.")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_index = {name: idx for idx, name in enumerate(header)}
    has_id_col = "QTRM ID" in col_index

    channels = _default_channels()
    for row_i, row in enumerate(rows[1:]):
        if has_id_col:
            id_col = col_index["QTRM ID"]
            if id_col >= len(row):
                continue
            qtrm_index = _clamp(row[id_col], 1, NUM_QTRM) - 1 if row[id_col] is not None else None
            if qtrm_index is None:
                continue
        else:
            qtrm_index = row_i
        if not (0 <= qtrm_index < NUM_QTRM):
            continue

        row_channels = []
        for ch in range(1, 5):
            values = {}
            for label, attr, lo, hi in _CHANNEL_FIELDS:
                col = col_index.get(f"Ch{ch} {label}")
                raw = row[col] if col is not None and col < len(row) else 0
                values[attr] = _clamp(raw if raw is not None else 0, lo, hi)
            row_channels.append(QTRMChannel(**values))
        channels[qtrm_index] = row_channels

    return channels


class DwellTab(QWidget):
    send_requested = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)

        content = QWidget()
        layout = QVBoxLayout(content)

        top_row = QHBoxLayout()
        self.import_btn = QPushButton("Import from Excel...")
        self.import_btn.clicked.connect(self._on_import_clicked)
        top_row.addWidget(self.import_btn)

        self.send_btn = QPushButton("Send Dwell")
        self.send_btn.clicked.connect(self.send_requested.emit)
        top_row.addWidget(self.send_btn)

        self.summary_label = QLabel("Not yet run")
        self.response_time_label = QLabel("")
        top_row.addWidget(self.summary_label)
        top_row.addWidget(self.response_time_label)
        top_row.addStretch(1)
        layout.addLayout(top_row)

        self.model = DwellTableModel()
        self.table = QTableView()
        self.table.setModel(self.model)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.table.setMinimumHeight(320)
        layout.addWidget(self.table)

        self.led_matrix = LedMatrix()
        layout.addWidget(self.led_matrix, 1)

        # Wrapped in a QScrollArea so this tab's minimumSizeHint stays small,
        # matching every other tab in this app.
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.NoFrame)
        scroll.setWidget(content)

        self.header_panel = HeaderPanel()

        outer = QHBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll, 1)
        outer.addWidget(self.header_panel)

    def _on_import_clicked(self):
        path, _ = QFileDialog.getOpenFileName(self, "Import Dwell Data", "", "Excel Files (*.xlsx)")
        if not path:
            return
        try:
            channels = load_channels_from_excel(path)
        except Exception as e:
            QMessageBox.warning(self, "Import failed", f"Could not read '{path}':\n{e}")
            return
        self.model.load_channels(channels)
        self.summary_label.setText("Imported from Excel - not yet sent")

    def get_channels(self):
        return self.model.get_channels()

    def mark_pending(self):
        self.summary_label.setText("Sent - waiting for response...")
        self.response_time_label.setText("")
        self.led_matrix.set_all(_PENDING_COLOR)

    def show_results(self, linked_flags):
        acked_count = sum(1 for v in linked_flags if v)
        self.summary_label.setText(f"{acked_count}/{NUM_QTRM} QTRMs acknowledged")
        self.led_matrix.set_results(linked_flags)

    def show_response_time(self, microseconds: float):
        self.response_time_label.setText(f"{microseconds:.0f} µs")

    def show_no_response(self):
        self.summary_label.setText("No response")
        self.response_time_label.setText("")
        self.led_matrix.set_all(_NOT_LINKED_COLOR)
